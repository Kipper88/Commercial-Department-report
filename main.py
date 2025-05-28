from utils.data.data_form import formatted_data, formatted_data_briefcase
from utils.data.data_loader import get_names, get_names_without_comm_departament
import const

import pandas as pd
from io import BytesIO
from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from pydantic import BaseModel

app = FastAPI()

from fastapi import HTTPException

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="templates"), name="static")

@app.get("/", response_class=HTMLResponse)
async def index():
    return """
    <a href="/commercial_departament">Коммерческий Департамент</a><br>
    <a href="/briefcase">Портфель</a><br>
    <a href="/briefcase_all_workers">Портфель (все сотрудники)</a><br>
    """

@app.get("/commercial_departament", response_class=HTMLResponse)
async def commercial_departament():
    return """
    <title>Коммерческий Департамент</title>
    <form method="post">   
        <label>Введите период дат в формате дд-мм-гггг,дд-мм-гггг через запятую (от, до)\n
        Примечание: лучше писать не больше недели, поскольку портал может
        возвращать ошибки из-за большого объема данных</label><br>
        <textarea name="dates_periods" rows="5" cols="50"></textarea><br>
        <button type="submit">Сгенерировать отчет</button>
    </form>
    """

@app.post("/commercial_departament", response_class=StreamingResponse)
async def handle_form(dates_periods: str = Form(...)):
    names = await get_names()

    init_workers_list = {key: 0 for key in names}
    
    dates_period = dates_periods
    return await generate_excel_report(names, init_workers_list, dates_period)

@app.get('/briefcase', response_class=HTMLResponse)
async def briefcase():
    return """
        <title>Портфель клиентов</title>
        <form method="post" action="/briefcase">   
            <button type="submit">Сгенерировать отчет</button>
        </form>
        <br>
        <p>ПРИМЕЧАНИЕ: отчет генерируется ~8 мин. Во время генерации не закрывайте страницу</p>
    """

@app.post('/briefcase', response_class=StreamingResponse)
async def handle_briefcase():
    names = await get_names()

    init_workers_list = {key: 0 for key in names}
    
    return await generate_briefcase_report(names, init_workers_list)

@app.get('/briefcase_all_workers', response_class=HTMLResponse)
async def briefcase_all_workers():
    
    return """
        <title>Портфель клиентов (все сотрудники)</title>
        <form method="post" action="/briefcase_all_workers">   
            <button type="submit">Сгенерировать отчет</button>
        </form>
        <br>
        <p>ПРИМЕЧАНИЕ: отчет генерируется ~8 мин. Во время генерации не закрывайте страницу</p>
    """
    
@app.post("/briefcase_all_workers", response_class=StreamingResponse)
async def handling_briefcase_all_workers():
    names = await get_names_without_comm_departament()
    init_workers_list = {key: 0 for key in names}
    
    return await generate_briefcase_report_without_comm_departament(names, init_workers_list)
        
class NamesRequest(BaseModel):
    names: list[str]

#
async def generate_excel_report(workers, init_workers_list, dates_period):
    try:
        margin, count_orders, framed_kp, meeting, communications, completed_tasks, plan_activity1, calls_2x_minutes = await formatted_data(init_workers_list, dates_period, workers)


        # Data with empty columns for spacing
        data = {
            "ФИО": ["Общий итог"] + workers,
            "Маржа": [margin[0]] + [margin[1][w] for w in workers],
            "": [""] * (len(workers) + 1),  # Empty column
            "Количество заказов": [count_orders[0]] + [count_orders[1][w] for w in workers],
            " ": [""] * (len(workers) + 1),  # Empty column
            "Количество оформленных КП": [framed_kp[0]] + [framed_kp[1][w] for w in workers],
            "  ": [""] * (len(workers) + 1),  # Empty column
            "Количество встреч": [meeting[0]] + [meeting[1][w] for w in workers],
            "   ": [""] * (len(workers) + 1),  # Empty column
            "Количество коммуникаций": [communications[0]] + [communications[1][w] for w in workers],
            "Количество выполненных задач": [completed_tasks[0]] + [completed_tasks[1][w] for w in workers],
            "План активности": [plan_activity1[0]] + [plan_activity1[1][w] for w in workers],
            "    ": [""] * (len(workers) + 1),  # Empty column
            "Исходящие вызовы": [calls_2x_minutes[0]] + [calls_2x_minutes[1][w] for w in workers],
        }

        df = pd.DataFrame(data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write headers two rows higher
            df.to_excel(writer, index=False, sheet_name='Отчет', startrow=2)

            worksheet = writer.sheets['Отчет']

            # Column widths with empty columns
            column_widths = {
                'A': 20,  # ФИО
                'B': 20,  # Маржа
                'C': 5,   # Empty column
                'D': 15,  # Количество заказов
                'E': 5,   # Empty column
                'F': 15,  # Количество оформленных КП
                'G': 5,   # Empty column
                'H': 15,  # Количество встреч
                'I': 5,   # Empty column
                'J': 15,  # Количество коммуникаций
                'K': 15,  # Количество выполненных задач
                'L': 15,  # План активности
                'M': 5,   # Empty column
                'N': 15,  # Исходящие вызовы
            }
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Define border styles
            full_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            no_right_border = Border(
                left=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            no_left_border = Border(
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            side_borders_only = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )

            # Apply borders, excluding first two rows, with custom borders for empty and adjacent columns
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=14):
                for cell in row:
                    if cell.column_letter in ['C', 'E', 'G', 'I', 'M']:  # Empty columns: only left and right borders
                        cell.border = side_borders_only
                    elif cell.column_letter in ['B', 'D', 'F', 'H', 'L']:  # Columns to the left of empty columns
                        cell.border = no_right_border
                    elif cell.column_letter in ['D', 'F', 'H', 'J', 'N']:  # Columns to the right of empty columns
                        cell.border = no_left_border
                    else:
                        cell.border = full_border

            # Style for "Общий итог" row (row 4)
            header_fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
            bold_font = Font(bold=True)
            for cell in worksheet[4]:  # "Общий итог" row
                if cell.column_letter not in ['A', 'C', 'E', 'G', 'I', 'M']:  # Skip ФИО and empty columns
                    cell.fill = header_fill
                    cell.font = bold_font

            # Blue accent 1 lighter 80% for headers (row 3)
            blue_fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
            header_font = Font(bold=True)
            for cell in worksheet[3]:  # Header row
                if cell.column_letter not in ['A', 'C', 'E', 'G', 'I', 'M']:  # Skip ФИО and empty columns
                    cell.fill = blue_fill
                    cell.font = header_font
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            # Format "Маржа" column as rubles, including "Общий итог" (B4:B...)
            for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):  # Start from row 4
                for cell in row:
                    cell.number_format = '#,##0.00 "₽"'

            # Conditional formatting for adjusted columns
            max_row = worksheet.max_row

            # Маржа (B)
            worksheet.conditional_formatting.add(f'B5:B{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['1500000'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'B5:B{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['750000'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'B5:B{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['500000'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'B5:B{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Количество заказов (D)
            worksheet.conditional_formatting.add(f'D5:D{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['23'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'D5:D{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['15'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'D5:D{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['8'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'D5:D{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Количество оформленных КП (F)
            worksheet.conditional_formatting.add(f'F5:F{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['2'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'F5:F{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['1'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'F5:F{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'F5:F{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Количество встреч (H)
            worksheet.conditional_formatting.add(f'H5:H{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['10'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'H5:H{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['7'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'H5:H{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['5'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'H5:H{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Количество коммуникаций (J)
            worksheet.conditional_formatting.add(f'J5:J{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['75'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'J5:J{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['50'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'J5:J{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['25'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'J5:J{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Количество выполненных задач (K)
            worksheet.conditional_formatting.add(f'K5:K{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['53'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'K5:K{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['35'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'K5:K{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['25'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'K5:K{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # План активности (L)
            worksheet.conditional_formatting.add(f'L5:L{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['35'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'L5:L{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['26'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'L5:L{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['7'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'L5:L{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Исходящие вызовы (N)
            worksheet.conditional_formatting.add(f'N5:N{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['62'],
                           fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
            worksheet.conditional_formatting.add(f'N5:N{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['40'],
                           fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')))
            worksheet.conditional_formatting.add(f'N5:N{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['25'],
                           fill=PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')))
            worksheet.conditional_formatting.add(f'N5:N{max_row}',
                CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                           fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')))

            # Add date period at the bottom
            max_row = worksheet.max_row
            worksheet[f'A{max_row + 2}'] = f"Период: {dates_period}"
            worksheet[f'A{max_row + 2}'].font = Font(bold=True)

        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=activity_report.xlsx"}
        )

    except:
        import traceback
        tb = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"Internal Server Error, please contact the developer. Possible reasons: date or names entered in the wrong format. {tb}")

async def generate_briefcase_report(workers, init_workers_list):
    try:
        try:
            __actually_first_order, __actually, __re_potencial, __potencial = await formatted_data_briefcase(workers, init_workers_list)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            return tb

        # Создание таблицы
        data = {
            "ФИО": ["Общий итог"] + workers,
            "Действующий. Первый заказ": [__actually_first_order[0]] + [__actually_first_order[1][w] for w in workers],
            "Действующий.": [__actually[0]] + [__actually[1][w] for w in workers],
            "Повторно потенциальный.": [__re_potencial[0]] + [__re_potencial[1][w] for w in workers],
            "Потенциальный": [__potencial[0]] + [__potencial[1][w] for w in workers],
        }

        df = pd.DataFrame(data)

        # Сохраняем в буфер
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Отчёт")
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=activity_report.xlsx"}
        )

    except Exception as e:
        # import traceback
        # tb = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"Internal Server Error, please contact the developer. Possible reasons: date or names entered in the wrong format.")
    
    
async def generate_briefcase_report_without_comm_departament(workers, init_workers_list):
    try:
        try:
            __actually_first_order, __actually, __re_potencial, __potencial = await formatted_data_briefcase(workers, init_workers_list)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            return tb

        # Создание таблицы
        data = {
            "ФИО": ["Общий итог"] + workers,
            "Действующий. Первый заказ": [__actually_first_order[0]] + [__actually_first_order[1][w] for w in workers],
            "Действующий.": [__actually[0]] + [__actually[1][w] for w in workers],
            "Повторно потенциальный.": [__re_potencial[0]] + [__re_potencial[1][w] for w in workers],
            "Потенциальный": [__potencial[0]] + [__potencial[1][w] for w in workers],
        }

        df = pd.DataFrame(data)

        # Сохраняем в буфер
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Отчёт")
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=activity_report.xlsx"}
        )

    except Exception as e:
        # import traceback
        # tb = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"Internal Server Error, please contact the developer. Possible reasons: date or names entered in the wrong format.")
